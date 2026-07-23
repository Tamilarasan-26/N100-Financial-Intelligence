import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE"
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFEB9C",
    end_color="FFEB9C"
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE"
)
BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD966",
    end_color="FFD966"
)

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATABASE_PATH = PROJECT_ROOT / "db" / "nifty100.db"

OUTPUT_FOLDER = PROJECT_ROOT / "output"

OUTPUT_FOLDER.mkdir(
    parents=True,
    exist_ok=True
)

OUTPUT_FILE = OUTPUT_FOLDER / "peer_comparison.xlsx"

def load_peer_groups():

    con = sqlite3.connect(DATABASE_PATH)

    df = pd.read_sql(
        "SELECT * FROM peer_groups",
        con
    )

    con.close()

    return df


def load_financial_ratios():

    con = sqlite3.connect(DATABASE_PATH)

    df = pd.read_sql(
        """
        SELECT *
        FROM financial_ratios
        WHERE period_type='ANNUAL'
        """,
        con
    )

    con.close()

    return df


def load_peer_percentiles():

    con = sqlite3.connect(DATABASE_PATH)

    df = pd.read_sql(
        "SELECT * FROM peer_percentiles",
        con
    )

    con.close()

    return df

def get_peer_group_names(peer_df):

    return sorted(
        peer_df["peer_group_name"]
        .dropna()
        .unique()
    )
    

def create_peer_group_dataframe(
    peer_df,
    financial_df,
    percentile_wide,
    peer_group
    ):

    companies = peer_df[
        peer_df["peer_group_name"] == peer_group
    ]

    merged_df = companies.merge(
        financial_df,
        on="company_id",
        how="left"
    )
    
    merged_df = merged_df.merge(
    percentile_wide,
    on="company_id",
    how="left",
    suffixes=("", "_Percentile")
    )


    return merged_df 

def pivot_percentiles(percentile_df):

    percentile_wide = percentile_df.pivot_table(
        index="company_id",
        columns="metric",
        values="percentile_rank"
    ).reset_index()

    return percentile_wide

def colour_percentile_columns(ws):

    percentile_columns = [
        "Asset Turnover",
        "Debt To Equity",
        "EPS CAGR 5Y",
        "Free Cash Flow",
        "Interest Coverage",
        "Net Profit Margin",
        "PAT CAGR 5Y",
        "ROCE",
        "ROE",
        "Revenue CAGR 5Y"
    ]

    header_map = {
        cell.value: cell.column
        for cell in ws[1]
    }

    for column_name in percentile_columns:

        if column_name not in header_map:
            continue

        col = header_map[column_name]

        for row in range(2, ws.max_row + 1):

            cell = ws.cell(row=row, column=col)

            if cell.value is None:
                continue

            cell.number_format = "0.00%"

            if cell.value >= 0.75:
                cell.fill = GREEN_FILL

            elif cell.value >= 0.25:
                cell.fill = YELLOW_FILL

            else:
                cell.fill = RED_FILL

def highlight_benchmark_row(ws):

    header_map = {
        cell.value: cell.column
        for cell in ws[1]
    }

    if "is_benchmark" not in header_map:
        return

    benchmark_col = header_map["is_benchmark"]

    for row in range(2, ws.max_row + 1):

        value = ws.cell(
            row=row,
            column=benchmark_col
        ).value

        if value in (1, True, "1", "True"):

            for col in range(1, ws.max_column + 1):

                ws.cell(
                    row=row,
                    column=col
                ).fill = BENCHMARK_FILL
                
def add_median_summary_row(ws):

    summary_row = ws.max_row + 1

    ws.cell(
        row=summary_row,
        column=1
    ).value = "Peer Group Median"

    for col in range(2, ws.max_column + 1):

        values = []

        for row in range(2, ws.max_row + 1):

            value = ws.cell(
                row=row,
                column=col
            ).value

            if isinstance(value, (int, float)):
                values.append(value)

        if values:

            values.sort()

            n = len(values)

            if n % 2 == 1:
                median = values[n // 2]
            else:
                median = (
                    values[n // 2 - 1] +
                    values[n // 2]
                ) / 2

            cell = ws.cell(
                row=summary_row,
                column=col
            )

            cell.value = median

peer_df = load_peer_groups()

financial_df = load_financial_ratios()

percentile_df = load_peer_percentiles()

percentile_wide = pivot_percentiles(percentile_df)

peer_groups = get_peer_group_names(peer_df)

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    for group in peer_groups:

        group_df = create_peer_group_dataframe(
            peer_df,
            financial_df,
            percentile_wide,
            group
        )
        
        group_df.to_excel(
            writer,
            sheet_name=group[:31],
            index=False
        )
        
workbook = load_workbook(OUTPUT_FILE)

for sheet in workbook.sheetnames:

    worksheet = workbook[sheet]

    add_median_summary_row(worksheet)

    colour_percentile_columns(worksheet)

    highlight_benchmark_row(worksheet)
    

workbook.save(OUTPUT_FILE)

print(f"Excel file created: {OUTPUT_FILE}")
